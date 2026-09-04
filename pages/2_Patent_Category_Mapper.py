import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font

from auth import require_password


# ============================================================
# PAGE CONFIGURATION
# ============================================================

st.set_page_config(
    page_title="Patent Category Mapper",
    page_icon="📊",
    layout="wide"
)

st.markdown(
    """
<style>
[data-testid="stSidebar"] {display: none;}
[data-testid="collapsedControl"] {display: none;}

.stApp {
    background-color: var(--background-color);
    color: var(--text-color);
}

.block-container {
    max-width: 1180px;
    padding-top: 4rem;
    padding-bottom: 3.5rem;
}

* {
    font-family: 'Inter', sans-serif;
}

.main-title {
    font-size: 2.35rem;
    font-weight: 700;
    color: #EE3C18;
    letter-spacing: -0.04em;
    line-height: 1.15;
    margin-bottom: 0.35rem;
}

.sub-title {
    font-size: 1.08rem;
    font-weight: 500;
    color: var(--text-color);
    margin-bottom: 1.05rem;
}

.block-container p {
    color: var(--text-color);
    opacity: 0.70;
    line-height: 1.65;
}

div[data-testid="stButton"] button {
    border-radius: 7px;
    font-size: 0.82rem;
    font-weight: 500;
    border: 1px solid rgba(128,128,128,0.25);
    background-color: var(--secondary-background-color);
    color: var(--text-color);
    min-height: 36px;
}

div[data-testid="stButton"] button:hover {
    border-color: #EE3C18;
    color: #EE3C18;
}

div[data-testid="stButton"] button[kind="primary"] {
    background-color: #EE3C18;
    border: 1px solid #EE3C18;
    color: #FFFFFF !important;
    font-weight: 600;
    min-height: 42px;
}

div[data-testid="stButton"] button[kind="primary"] p,
div[data-testid="stButton"] button[kind="primary"] span {
    color: #FFFFFF !important;
}

[data-testid="stFileUploader"] {
    background-color: var(--secondary-background-color);
    border: 1px dashed rgba(128,128,128,0.30);
    border-radius: 10px;
    padding: 0.8rem 1rem;
}

div[data-testid="stDownloadButton"] button {
    background-color: #EE3C18;
    border: 1px solid #EE3C18;
    color: #FFFFFF !important;
    border-radius: 8px;
    font-weight: 600;
    min-height: 42px;
}

div[data-testid="stDownloadButton"] button p,
div[data-testid="stDownloadButton"] button span {
    color: #FFFFFF !important;
}

hr {
    border-top: 1px solid rgba(128,128,128,0.22);
}
</style>
""",
    unsafe_allow_html=True
)

require_password()

if st.button("← Back to Home"):
    st.switch_page("Home.py")


# ============================================================
# PAGE HEADER
# ============================================================

st.markdown(
    '<div class="main-title">Patent Research</div>',
    unsafe_allow_html=True
)

st.markdown(
    '<div class="sub-title">Patent Category Mapper</div>',
    unsafe_allow_html=True
)

st.write(
    """
    Converts patent categorization information into individual
    category columns, allowing each patent to be clearly mapped
    to its applicable research categories.
    """
)

st.markdown(
    """
    <div style="
        margin-top: 1.25rem;
        margin-bottom: 2.2rem;
        padding: 1.15rem 1.5rem;
        background-color: var(--secondary-background-color);
        border: 1px solid rgba(128,128,128,0.20);
        border-radius: 10px;
    ">
        <div style="
            font-weight: 600;
            margin-bottom: 0.65rem;
            color: var(--text-color);
        ">
            Processing includes:
        </div>

        <ul style="
            margin: 0;
            padding-left: 1.35rem;
            color: var(--text-color);
            opacity: 0.70;
            line-height: 1.65;
        ">
            <li>Select the categorization column from the Excel headers</li>
            <li>Identify single or multiple categories assigned to each patent</li>
            <li>Use existing category columns when already present</li>
            <li>Create new category columns when categories are not already added</li>
            <li>Mark applicable patents with <strong>Y</strong></li>
            <li>Case-insensitive and whitespace-insensitive category matching</li>
            <li>Leave blank categorization cells unchanged</li>
            <li>Preserve all existing Excel data and columns</li>
            <li>Export the processed workbook as a new Excel file</li>
        </ul>
    </div>
    """,
    unsafe_allow_html=True
)


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def clean_value(value):
    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    return str(value).strip()


def normalize_for_match(value):
    """
    Case-insensitive + whitespace-insensitive matching.
    """
    return re.sub(
        r"\s+",
        "",
        clean_value(value)
    ).casefold()


def split_categories(value):
    """
    Categories can be separated by commas or line breaks.
    """

    value = clean_value(value)

    if not value:
        return []

    parts = re.split(
        r"[,;\r\n]+",
        value
    )

    result = []
    seen = set()

    for part in parts:

        category = re.sub(
            r"\s+",
            " ",
            part
        ).strip()

        if not category:
            continue

        key = normalize_for_match(category)

        if key not in seen:
            seen.add(key)
            result.append(category)

    return result


def process_workbook(
    file_bytes,
    categorization_column,
    categories_already_added
):

    wb = load_workbook(
        BytesIO(file_bytes)
    )

    ws = wb.active

    if ws.max_row < 2:
        raise ValueError(
            "The workbook does not contain data rows."
        )

    # --------------------------------------------------------
    # READ HEADERS
    # --------------------------------------------------------

    headers = [
        clean_value(
            ws.cell(1, col).value
        )
        for col in range(
            1,
            ws.max_column + 1
        )
    ]

    while headers and not headers[-1]:
        headers.pop()

    if categorization_column not in headers:
        raise ValueError(
            f'Categorization column "{categorization_column}" '
            "was not found."
        )

    category_col_index = (
        headers.index(categorization_column) + 1
    )

    # --------------------------------------------------------
    # EXTRACT CATEGORIES FROM EACH ROW
    # --------------------------------------------------------

    row_categories = {}
    all_categories = []

    for row in range(
        2,
        ws.max_row + 1
    ):

        categories = split_categories(
            ws.cell(
                row,
                category_col_index
            ).value
        )

        row_categories[row] = categories
        all_categories.extend(categories)

    # --------------------------------------------------------
    # UNIQUE CATEGORIES
    # --------------------------------------------------------

    category_lookup = {}

    for category in all_categories:

        key = normalize_for_match(category)

        if key and key not in category_lookup:
            category_lookup[key] = category

    unique_categories = list(
        category_lookup.values()
    )

    # --------------------------------------------------------
    # CATEGORY COLUMNS
    # --------------------------------------------------------

    category_columns = {}

    if categories_already_added:

        # Use existing columns only.
        # Matching = case-insensitive +
        # whitespace-insensitive.

        for col in range(
            1,
            ws.max_column + 1
        ):

            header = clean_value(
                ws.cell(1, col).value
            )

            if not header:
                continue

            category_columns[
                normalize_for_match(header)
            ] = col

        unmatched_categories = []

        for category in unique_categories:

            key = normalize_for_match(
                category
            )

            if key not in category_columns:
                unmatched_categories.append(
                    category
                )

    else:

        # ----------------------------------------------------
        # CREATE NEW CATEGORY COLUMNS AT END
        # ----------------------------------------------------

        start_col = ws.max_column + 1

        for offset, category in enumerate(
            unique_categories
        ):

            col_index = start_col + offset

            ws.cell(
                1,
                col_index,
                category
            )

            ws.cell(
                1,
                col_index
            ).font = Font(
                name="Arial",
                size=10,
                bold=True
            )

            category_columns[
                normalize_for_match(category)
            ] = col_index

        unmatched_categories = []

    # --------------------------------------------------------
    # MARK Y
    # --------------------------------------------------------

    y_count = 0

    for row, categories in row_categories.items():

        for category in categories:

            key = normalize_for_match(
                category
            )

            if key not in category_columns:
                continue

            col_index = category_columns[key]

            cell = ws.cell(
                row,
                col_index
            )

            cell.value = "Y"

            cell.font = Font(
                name="Arial",
                size=10
            )

            y_count += 1

    # --------------------------------------------------------
    # WORKSHEET FORMATTING
    # --------------------------------------------------------

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # --------------------------------------------------------
    # SAVE
    # --------------------------------------------------------

    output = BytesIO()

    wb.save(output)
    output.seek(0)

    stats = {
        "data_rows": ws.max_row - 1,
        "unique_categories": len(
            unique_categories
        ),
        "category_columns": len(
            category_columns
        ),
        "y_marks": y_count,
        "unmatched_categories":
            unmatched_categories
    }

    return output, stats


# ============================================================
# FILE UPLOAD
# ============================================================

st.markdown("### Upload Excel File")

uploaded_file = st.file_uploader(
    "Upload your Excel workbook",
    type=["xlsx"],
    label_visibility="collapsed"
)


# ============================================================
# INPUT OPTIONS
# ============================================================

if uploaded_file is not None:

    try:

        preview_bytes = (
            uploaded_file.getvalue()
        )

        preview_wb = load_workbook(
            BytesIO(preview_bytes),
            read_only=True
        )

        preview_ws = (
            preview_wb.active
        )

        preview_headers = [
            clean_value(
                preview_ws.cell(
                    1,
                    col
                ).value
            )
            for col in range(
                1,
                preview_ws.max_column + 1
            )
        ]

        preview_headers = [
            h for h in preview_headers
            if h
        ]

        preview_wb.close()

        if not preview_headers:
            st.error(
                "No column headers were found in row 1."
            )
            st.stop()

        # ----------------------------------------------------
        # CATEGORIZATION COLUMN
        # ----------------------------------------------------

        st.markdown(
            "### Categorization Column"
        )

        categorization_column = st.selectbox(
            "Select the column containing category information",
            options=preview_headers
        )

        # ----------------------------------------------------
        # EXISTING CATEGORY COLUMNS?
        # ----------------------------------------------------

        st.markdown(
            "### Category Columns"
        )

        categories_already_added = st.radio(
            "Are the category columns already present in the Excel file?",
            options=["Yes", "No"],
            horizontal=True,
            help=(
                "Yes = use existing category columns. "
                "No = create one column for each unique "
                "category at the end of the sheet."
            )
        )

        st.caption(
            "Categories can be separated by commas or "
            "line breaks. Matching is case-insensitive "
            "and whitespace-insensitive."
        )

    except Exception as e:

        st.error(
            f"Unable to read the workbook: {e}"
        )

        st.stop()


# ============================================================
# RUN BUTTON
# ============================================================

run_button = st.button(
    "RUN CATEGORY MAPPING",
    type="primary",
    use_container_width=True
)


# ============================================================
# RUN
# ============================================================

if run_button:

    if uploaded_file is None:
        st.error(
            "Please upload an Excel workbook first."
        )
        st.stop()

    try:

        with st.spinner(
            "Processing workbook..."
        ):

            output_buffer, stats = process_workbook(
                uploaded_file.getvalue(),
                categorization_column,
                categories_already_added == "Yes"
            )

        st.success(
            "Patent category mapping completed successfully."
        )

        # ----------------------------------------------------
        # SUMMARY
        # ----------------------------------------------------

        st.markdown(
            "### Processing Summary"
        )

        col1, col2, col3, col4 = st.columns(4)

        col1.metric(
            "Data Rows",
            f"{stats['data_rows']:,}"
        )

        col2.metric(
            "Unique Categories",
            f"{stats['unique_categories']:,}"
        )

        col3.metric(
            "Category Columns",
            f"{stats['category_columns']:,}"
        )

        col4.metric(
            "Y Marks",
            f"{stats['y_marks']:,}"
        )

        if stats["unmatched_categories"]:

            st.warning(
                f"{len(stats['unmatched_categories'])} "
                "category/categories in the "
                "categorization column did not have "
                "a matching existing category column "
                "and were therefore not marked."
            )

        # ----------------------------------------------------
        # DOWNLOAD
        # ----------------------------------------------------

        st.markdown("---")

        st.markdown(
            "### Download Result"
        )

        original_name = uploaded_file.name

        if original_name.lower().endswith(
            ".xlsx"
        ):

            download_name = (
                original_name[:-5]
                + "_CATEGORY_MAPPED.xlsx"
            )

        else:

            download_name = (
                original_name
                + "_CATEGORY_MAPPED.xlsx"
            )

        st.download_button(
            label="DOWNLOAD CATEGORY-MAPPED EXCEL",
            data=output_buffer,
            file_name=download_name,
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            type="primary",
            use_container_width=True
        )

    except Exception as e:

        st.error(
            "Processing failed."
        )

        st.exception(e)
