# ============================================================
# PATENT RESEARCH
# PARENT ASSIGNEE NORMALIZATION TOOL
# STREAMLIT CLOUD VERSION - WITH CHECKPOINTING
# ============================================================

import streamlit as st
import pandas as pd
import re
import time
import json
import difflib
import os
import hashlib

from collections import defaultdict, deque
from copy import copy
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font

from google import genai
from google.genai import types


# ============================================================
# 1. PAGE CONFIGURATION
# ============================================================

st.set_page_config(
    page_title="Parent Assignee Normalization",
    page_icon="🔎",
    layout="wide"
)


# ============================================================
# 2. INTERNAL CONFIGURATION
# ============================================================

CONFIG = {

    "ASSIGNEE_COLUMN": "ASSIGNEE",

    "PARENT_ASSIGNEE_COLUMN": "PARENT ASSIGNEE",

    "NORMALIZED_HEADER":
        "PARENT ASSIGNEE - NORMALIZED (AI)",

    "CONFIDENCE_HEADER":
        "PARENT ASSIGNEE - NORMALIZATION CONFIDENCE",

    "ULTIMATE_PARENT_HEADER":
        "PARENT ASSIGNEE - ULTIMATE PARENT (AI)",

    "ULTIMATE_PARENT_CONFIDENCE_HEADER":
        "PARENT ASSIGNEE - ULTIMATE PARENT CONFIDENCE",

    "ULTIMATE_PARENT_BATCH_SIZE": 25,

    "SIMILARITY_THRESHOLD": 0.85,

    # NOTE: verify this model name is currently valid for your keys.
    # If you get a 404 error, switch to "gemini-2.5-flash" or
    # "gemini-3.5-flash-lite" (both confirmed working earlier).
    "GEMINI_MODEL": "gemini-3.5-flash-lite",

    "GEMINI_RPM_PER_KEY": 8,

    "GEMINI_RPD_PER_KEY": 200,

    "GEMINI_MAX_CALLS": 300,

    "HEADER_ROW": 1,

    # Max seconds to wait for a free API key before failing fast
    # instead of hanging silently. Progress is checkpointed, so a
    # failure here just means: wait a bit, click Run again to resume.
    "KEY_WAIT_TIMEOUT_SECONDS": 45,
}


# ============================================================
# 3. GEMINI API KEYS
# ============================================================

def get_gemini_keys():

    raw_keys = ""

    try:
        raw_keys = st.secrets.get("GEMINI_KEYS", "")
    except Exception:
        raw_keys = ""

    if not raw_keys:
        raw_keys = os.environ.get("GEMINI_KEYS", "")

    keys = []

    for key in re.split(r"[,\n]+", raw_keys):

        key = key.strip()

        if key and key not in keys:

            keys.append(key)

    return keys


# ============================================================
# 4. CUSTOM CSS
# ============================================================

st.markdown(
    """
    <style>

    .main-title {
        font-size: 36px;
        font-weight: 700;
        margin-bottom: 0px;
    }

    .sub-title {
        font-size: 18px;
        color: #666666;
        margin-bottom: 25px;
    }

    .upload-box {
        padding: 20px;
        border-radius: 10px;
        border: 1px solid #DDDDDD;
        margin-top: 10px;
        margin-bottom: 20px;
    }

    </style>
    """,
    unsafe_allow_html=True
)


# ============================================================
# 5. PAGE TITLE
# ============================================================

st.markdown(
    '<div class="main-title">Patent Research</div>',
    unsafe_allow_html=True
)

st.markdown(
    '<div class="sub-title">Parent Assignee Normalization Tool</div>',
    unsafe_allow_html=True
)

st.write(
    """
    Normalize **PARENT ASSIGNEE** values using AI while
    keeping the original `ASSIGNEE` and `PARENT ASSIGNEE`
    columns unchanged.
    """
)

st.markdown(
    """
    **Processing includes:**

    - Parent Assignee normalization
    - Corporate entity resolution
    - Ultimate Parent Company identification
    - Confidence classification
    - Excel output generation
    - Joint venture / multi-entity cell detection (preserved, not merged)
    - Resumable progress - interrupted runs can continue instead of restarting
    """
)


# ============================================================
# 6. FILE UPLOAD
# ============================================================

st.markdown(
    "### Upload Excel File"
)

uploaded_file = st.file_uploader(
    "Upload your Excel workbook",
    type=["xlsx"],
    label_visibility="collapsed"
)


# ============================================================
# 6B. HEADER ROW SELECTION
# ============================================================

st.markdown("### Header Row")

st.warning(
    "Header cells must NOT be merged. Each column header should "
    "be a single, unmerged cell in one row. Merged header cells "
    "will cause incorrect column detection."
)

header_row_input = st.number_input(
    "Which row contains your column headers?",
    min_value=1,
    value=1,
    step=1
)

CONFIG["HEADER_ROW"] = int(header_row_input)


# ============================================================
# 7. HELPER FUNCTIONS
# ============================================================

def clean_value(value):

    if value is None:
        return ""

    try:

        if pd.isna(value):
            return ""

    except Exception:

        pass

    return str(value).strip()


def resolve_effective_parent(
    parent_value,
    assignee_value
):

    if (
        not parent_value
        or parent_value.strip()
        in ("-", "–", "—")
    ):

        return assignee_value

    return parent_value


def is_likely_joint_venture(value):
    """
    Detects values that likely represent two or more DISTINCT
    entities in one cell (e.g. a joint venture), based on a
    line break separating genuinely different names.

    A value with a line break where all lines are identical
    (or equivalent after basic text normalization) is treated
    as a duplicate, not a joint venture - it is not excluded
    from normalization.
    """

    if "\n" not in str(value):
        return False

    lines = [
        line.strip()
        for line in str(value).split("\n")
        if line.strip()
    ]

    if len(lines) < 2:
        return False

    normalized_lines = set(
        normalization_prep(line) for line in lines
    )

    return len(normalized_lines) > 1


def find_column(
    headers,
    target_name
):

    for header in headers:

        if (
            str(header).strip().upper()
            ==
            target_name.upper()
        ):

            return header

    target_clean = re.sub(
        r"\s+",
        " ",
        target_name.strip().lower()
    )

    for header in headers:

        header_clean = re.sub(
            r"\s+",
            " ",
            str(header).strip().lower()
        )

        if header_clean == target_clean:

            return header

    return None


def make_run_id(file_bytes, header_row, similarity_threshold, model):
    """
    Builds a stable identifier for this exact (file + settings)
    combination. Used as the checkpoint key in st.session_state so
    that resuming only happens when the same file/settings are
    used again - not accidentally across unrelated files.
    """

    file_hash = hashlib.md5(file_bytes).hexdigest()[:16]

    key_string = f"{file_hash}_{header_row}_{similarity_threshold}_{model}"

    return hashlib.md5(key_string.encode()).hexdigest()[:16]


def get_checkpoint(run_id):

    key = f"checkpoint_{run_id}"

    if key not in st.session_state:

        st.session_state[key] = {

            "normalized_map": {},

            "confidence_map": {},

            "gemini_calls": 0,

            "errors": 0,

            "next_cluster_index": 0,

            "next_fallback_index": 0,

            "ultimate_raw_parent_map": {},

            "ultimate_confidence_map": {},

            "next_ultimate_batch_index": 0,

            "stage": "clusters"

        }

    return st.session_state[key]


def save_checkpoint(run_id, checkpoint):

    st.session_state[f"checkpoint_{run_id}"] = checkpoint


def clear_checkpoint(run_id):

    key = f"checkpoint_{run_id}"

    if key in st.session_state:

        del st.session_state[key]


def has_checkpoint(run_id):

    key = f"checkpoint_{run_id}"

    if key not in st.session_state:
        return False

    checkpoint = st.session_state[key]

    return (
        checkpoint.get("gemini_calls", 0) > 0
        or checkpoint.get("next_cluster_index", 0) > 0
        or checkpoint.get("next_fallback_index", 0) > 0
    )


# ============================================================
# 8. TEXT NORMALIZATION
# ============================================================

def normalization_prep(value):

    value = str(value).strip().lower()

    value = re.sub(
        r",\s*\(?[a-z]{2}\)?\s*$",
        "",
        value
    )

    value = re.sub(
        r",.*$",
        "",
        value
    )

    value = re.sub(
        r"[,\.\(\)\[\]\{\}'\"]",
        " ",
        value
    )

    value = value.replace(
        "&",
        " and "
    )

    value = re.sub(
        r"[-_/]+",
        " ",
        value
    )

    value = re.sub(
        r"\s+",
        " ",
        value
    ).strip()

    return value


# ============================================================
# 9. FUZZY MATCHING
# ============================================================

def build_fuzzy_pairs(
    values,
    threshold
):

    pairs = []

    prepared = {
        value: normalization_prep(value)
        for value in values
    }

    for i, value_a in enumerate(values):

        for value_b in values[i + 1:]:

            a = prepared[value_a]

            b = prepared[value_b]

            if a == b:

                pairs.append(
                    (value_a, value_b)
                )

                continue

            similarity = difflib.SequenceMatcher(
                None,
                a,
                b
            ).ratio()

            if similarity >= threshold:

                pairs.append(
                    (value_a, value_b)
                )

    return pairs


# ============================================================
# 10. FUZZY CLUSTERS
# ============================================================

def build_clusters(pairs):

    parent = {}

    def find(x):

        parent.setdefault(
            x,
            x
        )

        while parent[x] != x:

            parent[x] = parent[parent[x]]

            x = parent[x]

        return x

    def union(x, y):

        root_x = find(x)

        root_y = find(y)

        if root_x != root_y:

            parent[root_x] = root_y

    for a, b in pairs:

        union(a, b)

    groups = defaultdict(set)

    for value in parent:

        root = find(value)

        groups[root].add(value)

    return [
        sorted(group)
        for group in groups.values()
        if len(group) > 1
    ]


# ============================================================
# 11. GEMINI KEY POOL
# ============================================================

class GeminiKeyPool:

    def __init__(
        self,
        keys,
        rpm_limit,
        rpd_limit
    ):

        self.keys = keys

        self.rpm_limit = rpm_limit

        self.rpd_limit = rpd_limit

        self.usage = {

            key: {

                "minute_hits": deque(),

                "day_count": 0,

                "day_reset":
                    time.time() + 86400,

                "cooldown_until": 0

            }

            for key in keys
        }


    def refresh(self, key):

        now = time.time()

        usage = self.usage[key]

        if now > usage["day_reset"]:

            usage["day_count"] = 0

            usage["day_reset"] = now + 86400

        while (
            usage["minute_hits"]
            and
            now
            -
            usage["minute_hits"][0]
            > 60
        ):

            usage["minute_hits"].popleft()


    def get_available_key(self):

        now = time.time()

        for key in self.keys:

            self.refresh(key)

            usage = self.usage[key]

            if now < usage["cooldown_until"]:

                continue

            if (
                len(usage["minute_hits"])
                >=
                self.rpm_limit
            ):

                continue

            if (
                usage["day_count"]
                >=
                self.rpd_limit
            ):

                continue

            return key

        return None


    def record_use(self, key):

        now = time.time()

        self.usage[key][
            "minute_hits"
        ].append(now)

        self.usage[key][
            "day_count"
        ] += 1


    def cooldown(
        self,
        key,
        seconds=65
    ):

        self.usage[key][
            "cooldown_until"
        ] = (
            time.time()
            + seconds
        )


    def wait_for_key(self, max_wait=45):

        waited = 0

        while waited < max_wait:

            key = self.get_available_key()

            if key:

                return key

            time.sleep(2)

            waited += 2

        return None


# ============================================================
# 12. GEMINI CLIENT
# ============================================================

def create_client(api_key):

    return genai.Client(
        api_key=api_key,
        http_options=types.HttpOptions(
            timeout=300000
        )
    )


def call_gemini(
    prompt,
    gemini_pool,
    model,
    max_retries=4,
    key_wait_timeout=45
):

    last_error = None

    for attempt in range(max_retries):

        key = gemini_pool.wait_for_key(max_wait=key_wait_timeout)

        if key is None:

            raise RuntimeError(
                "All Gemini API keys are currently rate-limited. "
                "Your progress so far has been saved - wait a "
                "minute and click Run again to resume."
            )

        try:

            client = create_client(key)

            response = client.models.generate_content(
                model=model,
                contents=[prompt],
                config=types.GenerateContentConfig(
                    temperature=0,
                    response_mime_type="application/json"
                )
            )

            gemini_pool.record_use(key)

            return json.loads(response.text)

        except Exception as e:

            err_msg = str(e).lower()

            if any(
                x in err_msg
                for x in [
                    "429",
                    "quota",
                    "resource_exhausted",
                    "rate limit"
                ]
            ):

                gemini_pool.cooldown(key)

                last_error = str(e)

                continue

            last_error = str(e)

            time.sleep(
                2 ** attempt
            )

    raise RuntimeError(
        "Gemini failed after retries: "
        + str(last_error)
    )


# ============================================================
# 13. PARENT CONTEXT
# ============================================================

def get_assignee_context(
    parent,
    parent_context
):

    values = sorted(
        parent_context.get(
            parent,
            set()
        )
    )

    return values[:30]


# ============================================================
# 14. GEMINI NORMALIZATION PROMPT
# ============================================================

def build_parent_prompt(
    parent_variants,
    context_by_parent
):

    variants_text = "\n".join(
        f"- {value}"
        for value in parent_variants
    )

    context_lines = []

    for parent in parent_variants:

        assignees = context_by_parent.get(
            parent,
            []
        )

        if assignees:

            context_lines.append(
                f"\nParent value: {parent}\n"
                f"Associated Assignees:\n"
                +
                "\n".join(
                    f"  - {a}"
                    for a in assignees
                )
            )

    context_text = (
        "\n".join(context_lines)
        if context_lines
        else
        "No Assignee context available."
    )

    prompt = f"""
You are an expert corporate entity-resolution analyst
working on a professional patent database.

The task is to normalize the PARENT ASSIGNEE.

Use the associated ASSIGNEE information as contextual
evidence when necessary.

Different Parent Assignee values may refer to the same
company because of spelling, punctuation, abbreviations,
legal suffixes or corporate renaming.

Do NOT merge unrelated companies.

Do NOT invent corporate relationships.

If evidence is insufficient, preserve the supplied
Parent Assignee rather than guessing.

Prefer the most complete formal corporate name.

The normalized result must represent ONE corporate entity.

------------------------------------------------------------
POTENTIAL PARENT ASSIGNEE VARIANTS
------------------------------------------------------------

{variants_text}

------------------------------------------------------------
ASSIGNEE CONTEXT
------------------------------------------------------------

{context_text}

------------------------------------------------------------
OUTPUT
------------------------------------------------------------

Return ONLY valid JSON.

{{
    "same_parent_entity": true,
    "normalized_parent_assignee": "Canonical Legal Name",
    "confidence": "High"
}}

If it cannot be confidently resolved:

{{
    "same_parent_entity": false,
    "normalized_parent_assignee": null,
    "confidence": "Low"
}}

Do not return markdown.
Do not return code fences.
"""

    return prompt


# ============================================================
# 15. ULTIMATE PARENT PROMPT
# ============================================================

def build_ownership_prompt(
    entity_batch,
    full_entity_list
):

    batch_text = "\n".join(
        f"- {e}"
        for e in entity_batch
    )

    full_text = "\n".join(
        f"- {e}"
        for e in full_entity_list
    )

    prompt = f"""
You are a corporate ownership research analyst.

You are given company names extracted from a patent database.

Some entities may be subsidiaries, divisions or business
units of OTHER companies that ALSO appear in the full list.

For each entity in ENTITIES TO EVALUATE, determine whether
it is currently controlled by another entity appearing in
the FULL ENTITY LIST.

Only report the parent if it appears in the FULL ENTITY LIST.

Never invent an external parent.

------------------------------------------------------------
RULES
------------------------------------------------------------

1. Identify CURRENT ownership only.

2. Do not use historical ownership if it is no longer current.

3. Do not treat spelling variants of the same company as
   parent/subsidiary relationships.

4. Do not report relationships merely because names are
   similar.

5. If no parent exists within the supplied list, omit it.

6. Only report relationships supported by strong evidence.

7. Do not report the relationship in both directions.

------------------------------------------------------------
FULL ENTITY LIST
------------------------------------------------------------

{full_text}

------------------------------------------------------------
ENTITIES TO EVALUATE
------------------------------------------------------------

{batch_text}

------------------------------------------------------------
OUTPUT
------------------------------------------------------------

Return ONLY valid JSON:

{{
    "relationships": [
        {{
            "subsidiary": "Exact name from ENTITIES TO EVALUATE",
            "ultimate_parent": "Exact name from FULL ENTITY LIST",
            "confidence": "High"
        }}
    ]
}}

If no relationships are found:

{{"relationships": []}}

Do not return markdown.
Do not return code fences.
"""

    return prompt


# ============================================================
# 16. PARSE WORKBOOK METADATA (cheap, always recomputed fresh)
# ============================================================

def parse_workbook_metadata(file_bytes, header_row):
    """
    Reads the uploaded file, extracts headers, unique parent
    values, joint-venture detection, and assignee context.
    This is fast and deterministic (no API calls), so it is
    always recomputed fresh rather than checkpointed - only
    the expensive Gemini resolution steps are checkpointed.
    """

    wb = load_workbook(BytesIO(file_bytes))

    if len(wb.sheetnames) != 1:

        raise ValueError(
            "This tool is designed for ONE sheet only. "
            f"Your workbook contains "
            f"{len(wb.sheetnames)} sheets."
        )

    ws = wb.worksheets[0]

    headers = []

    for cell in ws[header_row]:

        if cell.value is None:
            headers.append("")
        else:
            headers.append(str(cell.value).strip())

    assignee_column = find_column(headers, CONFIG["ASSIGNEE_COLUMN"])

    parent_assignee_column = find_column(
        headers, CONFIG["PARENT_ASSIGNEE_COLUMN"]
    )

    if assignee_column is None:
        raise ValueError("ASSIGNEE column was not found.")

    if parent_assignee_column is None:
        raise ValueError("PARENT ASSIGNEE column was not found.")

    data = list(ws.values)

    if len(data) <= header_row:
        raise ValueError("No data exists below the header row.")

    df = pd.DataFrame(data[header_row:], columns=headers)

    df[assignee_column] = df[assignee_column].apply(clean_value)

    df[parent_assignee_column] = df[parent_assignee_column].apply(clean_value)

    df["EFFECTIVE_PARENT"] = df.apply(
        lambda row: resolve_effective_parent(
            row[parent_assignee_column],
            row[assignee_column]
        ),
        axis=1
    )

    unique_parents = sorted(
        set(v for v in df["EFFECTIVE_PARENT"] if v)
    )

    jv_values = set(
        v for v in unique_parents if is_likely_joint_venture(v)
    )

    normal_values = [
        v for v in unique_parents if v not in jv_values
    ]

    parent_context = defaultdict(set)

    for _, row in df.iterrows():

        parent = row["EFFECTIVE_PARENT"]

        assignee = clean_value(row[assignee_column])

        if parent and assignee:

            parent_context[parent].add(assignee)

    data_rows = max(0, ws.max_row - header_row)

    # Free the DataFrame - not needed again. The output-writing
    # step reads directly from the worksheet, not from df.
    del df
    del data
    del wb

    fuzzy_pairs = build_fuzzy_pairs(
        normal_values, CONFIG["SIMILARITY_THRESHOLD"]
    )

    clusters = build_clusters(fuzzy_pairs)

    return {
        "assignee_column": assignee_column,
        "parent_assignee_column": parent_assignee_column,
        "unique_parents": unique_parents,
        "jv_values": jv_values,
        "normal_values": normal_values,
        "parent_context": parent_context,
        "fuzzy_pairs": fuzzy_pairs,
        "clusters": clusters,
        "data_rows": data_rows,
    }


# ============================================================
# 17. RESOLVE NORMALIZATION (checkpointed - the expensive part)
# ============================================================

def resolve_parent_normalization(file_bytes, gemini_keys, run_id):

    model = CONFIG["GEMINI_MODEL"]
    max_calls = CONFIG["GEMINI_MAX_CALLS"]
    key_wait_timeout = CONFIG["KEY_WAIT_TIMEOUT_SECONDS"]

    meta = parse_workbook_metadata(file_bytes, CONFIG["HEADER_ROW"])

    normal_values = meta["normal_values"]
    clusters = meta["clusters"]
    parent_context = meta["parent_context"]

    checkpoint = get_checkpoint(run_id)

    normalized_map = checkpoint["normalized_map"]
    confidence_map = checkpoint["confidence_map"]
    gemini_calls = checkpoint["gemini_calls"]
    errors = checkpoint["errors"]

    gemini_pool = GeminiKeyPool(
        gemini_keys,
        CONFIG["GEMINI_RPM_PER_KEY"],
        CONFIG["GEMINI_RPD_PER_KEY"]
    )

    progress_bar = st.progress(0, text="Normalizing Parent Assignee values...")
    status_placeholder = st.empty()

    total_clusters = len(clusters)

    # --------------------------------------------------------
    # Stage: clusters (resumes from checkpoint["next_cluster_index"])
    # --------------------------------------------------------

    if checkpoint["stage"] == "clusters":

        for cluster_index in range(checkpoint["next_cluster_index"], total_clusters):

            if gemini_calls >= max_calls:
                break

            cluster = clusters[cluster_index]

            status_placeholder.caption(f"Checking: {cluster[0][:70]}")

            context = {
                p: get_assignee_context(p, parent_context) for p in cluster
            }

            prompt = build_parent_prompt(cluster, context)

            result = call_gemini(
                prompt, gemini_pool, model, key_wait_timeout=key_wait_timeout
            )

            gemini_calls += 1

            normalized = result.get("normalized_parent_assignee")
            confidence = result.get("confidence", "Unknown")
            same_entity = result.get("same_parent_entity", False)

            if same_entity and normalized:

                normalized = str(normalized).strip()

                for original in cluster:
                    normalized_map[original] = normalized
                    confidence_map[original] = confidence

            # Checkpoint immediately after every successful call.
            checkpoint["normalized_map"] = normalized_map
            checkpoint["confidence_map"] = confidence_map
            checkpoint["gemini_calls"] = gemini_calls
            checkpoint["next_cluster_index"] = cluster_index + 1
            save_checkpoint(run_id, checkpoint)

            progress_bar.progress(
                min((cluster_index + 1) / max(total_clusters, 1), 1.0),
                text=f"Normalizing Parent Assignee values... {cluster_index + 1}/{max(total_clusters,1)}"
            )

        checkpoint["stage"] = "fallbacks"
        save_checkpoint(run_id, checkpoint)

    status_placeholder.empty()

    # --------------------------------------------------------
    # Stage: unclustered fallbacks
    # --------------------------------------------------------

    clustered_values = set()
    for cluster in clusters:
        clustered_values.update(cluster)

    unclustered_fallbacks = [
        v for v in normal_values
        if v not in normalized_map and v not in clustered_values
    ]

    if checkpoint["stage"] == "fallbacks":

        canonical_names_so_far = sorted(set(normalized_map.values()))

        for fb_index in range(checkpoint["next_fallback_index"], len(unclustered_fallbacks)):

            if gemini_calls >= max_calls:
                break

            value = unclustered_fallbacks[fb_index]

            context = {value: get_assignee_context(value, parent_context)}
            prompt = build_parent_prompt([value], context)

            if canonical_names_so_far:
                prompt += (
                    "\n\nKNOWN CANONICAL PARENT ENTITIES:\n"
                    + "\n".join(f"- {c}" for c in canonical_names_so_far)
                )

            result = call_gemini(
                prompt, gemini_pool, model, key_wait_timeout=key_wait_timeout
            )

            gemini_calls += 1

            if result.get("same_parent_entity") and result.get("normalized_parent_assignee"):

                normalized = str(result["normalized_parent_assignee"]).strip()
                normalized_map[value] = normalized
                confidence_map[value] = result.get("confidence", "Unknown")
                canonical_names_so_far.append(normalized)

            checkpoint["normalized_map"] = normalized_map
            checkpoint["confidence_map"] = confidence_map
            checkpoint["gemini_calls"] = gemini_calls
            checkpoint["next_fallback_index"] = fb_index + 1
            save_checkpoint(run_id, checkpoint)

        checkpoint["stage"] = "reconciliation"
        save_checkpoint(run_id, checkpoint)

    # --------------------------------------------------------
    # Stage: exact-match + canonical reconciliation (cheap, no API)
    # --------------------------------------------------------

    prepared_to_value = {}

    for value in normal_values:

        prepared = normalization_prep(value)

        if prepared in prepared_to_value:

            existing = prepared_to_value[prepared]

            if value not in normalized_map:
                normalized_map[value] = normalized_map.get(existing, existing)
                confidence_map[value] = "High"

        else:
            prepared_to_value[prepared] = value

    canonical_names = sorted(set(normalized_map.values()))

    for value in normal_values:

        if value in normalized_map:
            continue

        prepped_value = normalization_prep(value)

        best_match = None
        best_score = 0.0

        for canonical in canonical_names:
            score = difflib.SequenceMatcher(
                None, prepped_value, normalization_prep(canonical)
            ).ratio()
            if score > best_score:
                best_score = score
                best_match = canonical

        if best_match and best_score >= 0.75:
            normalized_map[value] = best_match
            confidence_map[value] = "Medium"

    checkpoint["normalized_map"] = normalized_map
    checkpoint["confidence_map"] = confidence_map
    checkpoint["stage"] = "ultimate"
    save_checkpoint(run_id, checkpoint)

    # --------------------------------------------------------
    # Stage: ultimate parent resolution (resumable by batch)
    # --------------------------------------------------------

    canonical_names_full = sorted(set(normalized_map.values()))
    batch_size = CONFIG["ULTIMATE_PARENT_BATCH_SIZE"]

    raw_parent_map = checkpoint["ultimate_raw_parent_map"]
    confidence_map_up = checkpoint["ultimate_confidence_map"]

    total_batches = max(1, (len(canonical_names_full) + batch_size - 1) // batch_size)

    up_progress = st.progress(0, text="Resolving ultimate parent companies...")

    if checkpoint["stage"] == "ultimate":

        start_index = checkpoint["next_ultimate_batch_index"] * batch_size

        for batch_num, i in enumerate(
            range(start_index, len(canonical_names_full), batch_size),
            start=checkpoint["next_ultimate_batch_index"] + 1
        ):

            if gemini_calls >= max_calls:
                break

            batch = canonical_names_full[i:i + batch_size]

            prompt = build_ownership_prompt(batch, canonical_names_full)

            try:
                result = call_gemini(
                    prompt, gemini_pool, model, key_wait_timeout=key_wait_timeout
                )
                gemini_calls += 1
            except Exception as e:
                st.warning(f"Ultimate parent batch error: {e}")
                checkpoint["next_ultimate_batch_index"] = batch_num
                save_checkpoint(run_id, checkpoint)
                continue

            for rel in result.get("relationships", []):

                subsidiary = str(rel.get("subsidiary", "")).strip()
                parent = str(rel.get("ultimate_parent", "")).strip()
                confidence = rel.get("confidence", "Unknown")

                if not subsidiary or not parent:
                    continue
                if subsidiary not in canonical_names_full:
                    continue
                if parent not in canonical_names_full:
                    continue
                if subsidiary == parent:
                    continue

                raw_parent_map[subsidiary] = parent
                confidence_map_up[subsidiary] = confidence

            checkpoint["ultimate_raw_parent_map"] = raw_parent_map
            checkpoint["ultimate_confidence_map"] = confidence_map_up
            checkpoint["gemini_calls"] = gemini_calls
            checkpoint["next_ultimate_batch_index"] = batch_num
            save_checkpoint(run_id, checkpoint)

            up_progress.progress(
                min(batch_num / total_batches, 1.0),
                text=f"Resolving ultimate parent companies... {batch_num}/{total_batches}"
            )

        checkpoint["stage"] = "done"
        save_checkpoint(run_id, checkpoint)

    up_progress.empty()
    progress_bar.empty()

    def resolve_chain(name, visited=None):

        if visited is None:
            visited = set()
        if name in visited:
            return name
        visited.add(name)
        if name in raw_parent_map:
            return resolve_chain(raw_parent_map[name], visited)
        return name

    ultimate_parent_map = {}

    for name in canonical_names_full:
        resolved = resolve_chain(name)
        if resolved != name:
            ultimate_parent_map[name] = resolved

    return {
        "meta": meta,
        "normalized_map": normalized_map,
        "confidence_map": confidence_map,
        "ultimate_parent_map": ultimate_parent_map,
        "ultimate_confidence_map": confidence_map_up,
        "gemini_calls": gemini_calls,
        "errors": errors,
    }


# ============================================================
# 18. BUILD OUTPUT WORKBOOK
# ============================================================

def build_output_workbook(file_bytes, resolution, input_filename, partial=False):

    header_row = CONFIG["HEADER_ROW"]

    meta = resolution["meta"]
    assignee_column = meta["assignee_column"]
    parent_assignee_column = meta["parent_assignee_column"]
    unique_parents = meta["unique_parents"]
    jv_values = meta["jv_values"]
    parent_context = meta["parent_context"]
    fuzzy_pairs = meta["fuzzy_pairs"]
    clusters = meta["clusters"]
    data_rows = meta["data_rows"]

    normalized_map = resolution["normalized_map"]
    confidence_map = resolution["confidence_map"]
    ultimate_parent_map = resolution["ultimate_parent_map"]
    ultimate_confidence_map = resolution["ultimate_confidence_map"]
    gemini_calls = resolution["gemini_calls"]
    errors = resolution["errors"]

    wb = load_workbook(BytesIO(file_bytes))
    ws = wb.worksheets[0]

    output_headers = [
        CONFIG["NORMALIZED_HEADER"],
        CONFIG["CONFIDENCE_HEADER"],
        CONFIG["ULTIMATE_PARENT_HEADER"],
        CONFIG["ULTIMATE_PARENT_CONFIDENCE_HEADER"]
    ]

    existing_headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in ws[header_row]
    ]

    for output_header in output_headers:
        if output_header in existing_headers:
            col_index = existing_headers.index(output_header) + 1
            ws.delete_cols(col_index, 1)
            existing_headers.pop(col_index - 1)

    output_start_col = ws.max_column + 1

    for offset, header in enumerate(output_headers):
        ws.cell(header_row, output_start_col + offset, header)

    header_font = Font(name="Arial", size=10, bold=True)

    for col in range(output_start_col, output_start_col + 4):
        ws.cell(header_row, col).font = copy(header_font)

    assignee_col = None
    parent_col = None

    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if value is None:
            continue
        value = str(value).strip()
        if value.upper() == CONFIG["ASSIGNEE_COLUMN"].upper():
            assignee_col = col
        if value.upper() == CONFIG["PARENT_ASSIGNEE_COLUMN"].upper():
            parent_col = col

    if assignee_col is None:
        raise RuntimeError("ASSIGNEE column could not be located.")
    if parent_col is None:
        raise RuntimeError("PARENT ASSIGNEE column could not be located.")

    normalized_rows = 0
    unchanged_rows = 0
    blank_parent_rows = 0
    jv_preserved_rows = 0

    for row in range(header_row + 1, ws.max_row + 1):

        original_parent = clean_value(ws.cell(row, parent_col).value)
        original_assignee = clean_value(ws.cell(row, assignee_col).value)
        effective_parent = resolve_effective_parent(original_parent, original_assignee)

        if not effective_parent:
            blank_parent_rows += 1
            continue

        if effective_parent in jv_values:
            normalized_value = effective_parent
            confidence = "Preserved (Possible Joint Venture)"
            jv_preserved_rows += 1
        elif effective_parent in normalized_map:
            normalized_value = normalized_map[effective_parent]
            confidence = confidence_map.get(effective_parent, "Unknown")
            normalized_rows += 1
        else:
            normalized_value = effective_parent
            confidence = "Unchanged"
            unchanged_rows += 1

        if effective_parent in jv_values:
            ultimate_value = normalized_value
            ultimate_confidence = "Preserved (Possible Joint Venture)"
        else:
            ultimate_value = ultimate_parent_map.get(normalized_value, normalized_value)
            if normalized_value in ultimate_parent_map:
                ultimate_confidence = ultimate_confidence_map.get(normalized_value, "Unknown")
            else:
                ultimate_confidence = "Same as Normalized"

        ws.cell(row, output_start_col, normalized_value)
        ws.cell(row, output_start_col + 1, confidence)
        ws.cell(row, output_start_col + 2, ultimate_value)
        ws.cell(row, output_start_col + 3, ultimate_confidence)

    for row in range(header_row + 1, ws.max_row + 1):
        for col in range(output_start_col, output_start_col + 4):
            cell = ws.cell(row, col)
            if cell.value is not None:
                new_font = copy(cell.font)
                new_font.name = "Arial"
                new_font.sz = 10
                cell.font = new_font

    # ---- Parent Normalization Map sheet ----

    if "Parent Normalization Map" in wb.sheetnames:
        del wb["Parent Normalization Map"]

    map_ws = wb.create_sheet("Parent Normalization Map")

    map_ws.append([
        "Original Parent Assignee", "Normalized Parent Assignee", "Confidence",
        "Ultimate Parent Company", "Ultimate Parent Confidence",
        "Associated Assignee Count", "Associated Assignees"
    ])

    for parent in unique_parents:

        if parent in jv_values:
            normalized = parent
            confidence = "Preserved (Possible Joint Venture)"
            ultimate = parent
            ultimate_confidence = "Preserved (Possible Joint Venture)"
        elif parent in normalized_map:
            normalized = normalized_map[parent]
            confidence = confidence_map.get(parent, "Unknown")
            ultimate = ultimate_parent_map.get(normalized, normalized)
            ultimate_confidence = ultimate_confidence_map.get(normalized, "Same as Normalized")
        else:
            normalized = parent
            confidence = "Unchanged"
            ultimate = ultimate_parent_map.get(normalized, normalized)
            ultimate_confidence = ultimate_confidence_map.get(normalized, "Same as Normalized")

        associated_assignees = sorted(parent_context.get(parent, set()))

        map_ws.append([
            parent, normalized, confidence, ultimate, ultimate_confidence,
            len(associated_assignees), "; ".join(associated_assignees)
        ])

    for cell in map_ws[1]:
        cell.font = Font(name="Arial", size=10, bold=True)

    for row in map_ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)

    map_ws.freeze_panes = "A2"
    map_ws.auto_filter.ref = f"A1:G{map_ws.max_row}"

    map_ws.column_dimensions["A"].width = 45
    map_ws.column_dimensions["B"].width = 45
    map_ws.column_dimensions["C"].width = 18
    map_ws.column_dimensions["D"].width = 45
    map_ws.column_dimensions["E"].width = 25
    map_ws.column_dimensions["F"].width = 22
    map_ws.column_dimensions["G"].width = 80

    # ---- Summary sheet ----

    if "Parent Normalization Summary" in wb.sheetnames:
        del wb["Parent Normalization Summary"]

    summary_ws = wb.create_sheet("Parent Normalization Summary")
    summary_ws.append(["Metric", "Value"])

    summary_rows = [
        ["Input File", input_filename],
        ["Sheet", ws.title],
        ["Run Status", "PARTIAL - interrupted before completion" if partial else "Complete"],
        ["Assignee Column", assignee_column],
        ["Parent Assignee Column", parent_assignee_column],
        ["Data Rows", data_rows],
        ["Unique Parent Assignee Values", len(unique_parents)],
        ["Possible Joint Venture Values (Preserved)", len(jv_values)],
        ["Fuzzy Candidate Pairs", len(fuzzy_pairs)],
        ["Fuzzy Candidate Clusters", len(clusters)],
        ["Gemini Calls", gemini_calls],
        ["Gemini Errors", errors],
        ["Rows With Normalized Parent", normalized_rows],
        ["Rows Left Unchanged", unchanged_rows],
        ["Joint Venture Rows Preserved", jv_preserved_rows],
        ["Blank Parent Assignee Rows", blank_parent_rows],
        ["Ultimate Parent Relationships", len(ultimate_parent_map)],
    ]

    for r in summary_rows:
        summary_ws.append(r)

    for cell in summary_ws[1]:
        cell.font = Font(name="Arial", size=10, bold=True)

    for row in summary_ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)

    summary_ws.column_dimensions["A"].width = 40
    summary_ws.column_dimensions["B"].width = 35

    output_buffer = BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)

    stats = {
        "data_rows": data_rows,
        "unique_parents": len(unique_parents),
        "jv_values": len(jv_values),
        "fuzzy_pairs": len(fuzzy_pairs),
        "clusters": len(clusters),
        "gemini_calls": gemini_calls,
        "errors": errors,
        "normalized_rows": normalized_rows,
        "unchanged_rows": unchanged_rows,
        "jv_preserved_rows": jv_preserved_rows,
        "blank_rows": blank_parent_rows,
        "ultimate_relationships": len(ultimate_parent_map),
    }

    return output_buffer, stats


# ============================================================
# 19. RUN BUTTON
# ============================================================

run_button = st.button(
    "RUN PARENT ASSIGNEE NORMALIZATION",
    type="primary",
    use_container_width=True
)

if uploaded_file is not None:

    _file_bytes_preview = uploaded_file.getvalue()
    _run_id_preview = make_run_id(
        _file_bytes_preview,
        CONFIG["HEADER_ROW"],
        CONFIG["SIMILARITY_THRESHOLD"],
        CONFIG["GEMINI_MODEL"]
    )

    if has_checkpoint(_run_id_preview):

        cp = get_checkpoint(_run_id_preview)

        st.info(
            f"An interrupted run was found for this file/settings "
            f"(stage: {cp['stage']}, {cp['gemini_calls']} calls already made). "
            f"Clicking Run will **resume** from where it left off, not restart."
        )

        if st.button("Discard saved progress and start fresh instead"):
            clear_checkpoint(_run_id_preview)
            st.rerun()


# ============================================================
# 20. RUN
# ============================================================

if run_button:

    if uploaded_file is None:
        st.error("Please upload an Excel workbook first.")
        st.stop()

    gemini_keys = get_gemini_keys()

    if not gemini_keys:
        st.error("Gemini API keys are not configured for this app.")
        st.info(
            "Make sure a secret named GEMINI_KEYS is set under "
            "this app's Settings -> Secrets on Streamlit Cloud, "
            "formatted as: GEMINI_KEYS = \"key1,key2,key3\""
        )
        st.stop()

    file_bytes = uploaded_file.getvalue()

    run_id = make_run_id(
        file_bytes,
        CONFIG["HEADER_ROW"],
        CONFIG["SIMILARITY_THRESHOLD"],
        CONFIG["GEMINI_MODEL"]
    )

    st.markdown("---")
    st.subheader("Processing")
    st.write(f"**Input file:** {uploaded_file.name}")

    try:

        with st.spinner("Processing workbook..."):

            resolution = resolve_parent_normalization(file_bytes, gemini_keys, run_id)

            output_buffer, stats = build_output_workbook(
                file_bytes, resolution, uploaded_file.name, partial=False
            )

        # Success - safe to discard the checkpoint now.
        clear_checkpoint(run_id)

        st.success("Parent Assignee normalization completed successfully.")

        st.subheader("Processing Summary")

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Data Rows", f"{stats['data_rows']:,}")
        col2.metric("Unique Parents", f"{stats['unique_parents']:,}")
        col3.metric("Normalized Rows", f"{stats['normalized_rows']:,}")
        col4.metric("Gemini Calls", f"{stats['gemini_calls']:,}")

        col5, col6, col7, col8 = st.columns(4)
        col5.metric("Fuzzy Pairs", f"{stats['fuzzy_pairs']:,}")
        col6.metric("Fuzzy Clusters", f"{stats['clusters']:,}")
        col7.metric("Unchanged Rows", f"{stats['unchanged_rows']:,}")
        col8.metric("Ultimate Parents", f"{stats['ultimate_relationships']:,}")

        if stats.get("jv_preserved_rows", 0) > 0:
            st.info(
                f"{stats['jv_preserved_rows']} row(s) appear to contain "
                "multiple entities in one cell (possible joint ventures) "
                "and were preserved as-is, not merged."
            )

        if stats["errors"] > 0:
            st.warning(
                f"{stats['errors']} Gemini processing errors occurred. "
                "Those values were left unchanged where no confident "
                "normalization was available."
            )

        st.markdown("---")
        st.subheader("Download Result")

        original_name = uploaded_file.name
        if original_name.lower().endswith(".xlsx"):
            download_name = original_name[:-5] + "_PARENT_ASSIGNEE_NORMALIZED.xlsx"
        else:
            download_name = original_name + "_PARENT_ASSIGNEE_NORMALIZED.xlsx"

        st.download_button(
            label="DOWNLOAD NORMALIZED EXCEL",
            data=output_buffer,
            file_name=download_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )

        st.markdown("---")
        st.subheader("Output Columns Added")
        st.success("PARENT ASSIGNEE - NORMALIZED (AI)")
        st.success("PARENT ASSIGNEE - NORMALIZATION CONFIDENCE")
        st.success("PARENT ASSIGNEE - ULTIMATE PARENT (AI)")
        st.success("PARENT ASSIGNEE - ULTIMATE PARENT CONFIDENCE")
        st.info("Original ASSIGNEE and PARENT ASSIGNEE columns are not modified.")
        st.info(
            "Cells containing multiple entities (e.g. joint ventures, "
            "detected via line breaks) are preserved as-is and never "
            "merged into a single normalized name."
        )
        st.info("No Reason column or Reason field is created.")

    except Exception as e:

        st.error(
            "Processing was interrupted before completion. "
            "Your progress up to this point has been saved."
        )
        st.exception(e)

        # Attempt to build a partial output from whatever was
        # resolved before the interruption, using the checkpoint.
        try:

            checkpoint = get_checkpoint(run_id)

            if checkpoint["gemini_calls"] > 0 or checkpoint["normalized_map"]:

                meta = parse_workbook_metadata(file_bytes, CONFIG["HEADER_ROW"])

                partial_resolution = {
                    "meta": meta,
                    "normalized_map": checkpoint["normalized_map"],
                    "confidence_map": checkpoint["confidence_map"],
                    "ultimate_parent_map": {},
                    "ultimate_confidence_map": {},
                    "gemini_calls": checkpoint["gemini_calls"],
                    "errors": checkpoint["errors"],
                }

                # Resolve ultimate parents from whatever was saved, if any.
                raw_parent_map = checkpoint.get("ultimate_raw_parent_map", {})
                confidence_map_up = checkpoint.get("ultimate_confidence_map", {})

                def _resolve_chain(name, visited=None):
                    if visited is None:
                        visited = set()
                    if name in visited:
                        return name
                    visited.add(name)
                    if name in raw_parent_map:
                        return _resolve_chain(raw_parent_map[name], visited)
                    return name

                canonical_names_full = sorted(set(checkpoint["normalized_map"].values()))
                ultimate_map = {}
                for name in canonical_names_full:
                    resolved = _resolve_chain(name)
                    if resolved != name:
                        ultimate_map[name] = resolved

                partial_resolution["ultimate_parent_map"] = ultimate_map
                partial_resolution["ultimate_confidence_map"] = confidence_map_up

                partial_buffer, partial_stats = build_output_workbook(
                    file_bytes, partial_resolution, uploaded_file.name, partial=True
                )

                st.warning(
                    f"Partial results are available: {partial_stats['normalized_rows']:,} "
                    f"row(s) were normalized before the interruption. Rows not yet "
                    f"processed are marked 'Unchanged'. You can download this partial "
                    f"file now, or click Run again to resume and complete processing."
                )

                original_name = uploaded_file.name
                if original_name.lower().endswith(".xlsx"):
                    partial_name = original_name[:-5] + "_PARENT_ASSIGNEE_PARTIAL.xlsx"
                else:
                    partial_name = original_name + "_PARENT_ASSIGNEE_PARTIAL.xlsx"

                st.download_button(
                    label="DOWNLOAD PARTIAL RESULTS",
                    data=partial_buffer,
                    file_name=partial_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            else:

                st.info(
                    "No progress had been made yet, so there is no partial "
                    "file to offer. Click Run again to retry."
                )

        except Exception as inner_e:

            st.error(
                "Could not build a partial results file either. "
                "Your checkpoint progress is still saved internally - "
                "click Run again to resume."
            )
            st.exception(inner_e)
